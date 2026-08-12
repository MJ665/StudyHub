"""batch_reports endpoints (moved verbatim from modules/reporting/routers/cohort_reports.py)."""
from fastapi import APIRouter

from modules.reporting.routers.cohort_shared import *  # noqa: F401,F403

router = APIRouter()

@router.get("/batch/{batch_id}/summary")
async def get_batch_report(
    batch_id: int,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(require_ldadmin),
):
    await db.run_sync(lambda s: assert_batch_in_org(batch_id, s, current_user))
    batch = await db.run_sync(lambda s: s.query(models.Batch).filter(models.Batch.id == batch_id).first())
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")

    groups = await db.run_sync(lambda s: s.query(models.Group).filter(models.Group.batch_id == batch_id).all())
    group_ids = [g.id for g in groups]

    # Date filtering logic. The query is built conditionally, so it is assembled
    # AND executed inside one run_sync call — a lambda cannot hold the branches.
    def _load_attempts(sync_db):
        q = (
            sync_db.query(models.Attempt, models.User)
            .join(models.User, models.Attempt.user_id == models.User.id)
            .filter(models.User.group_id.in_(group_ids))
        )
        if start_date:
            q = q.filter(
                models.Attempt.attempted_at
                >= datetime.datetime.fromisoformat(start_date)
            )
        if end_date:
            q = q.filter(
                models.Attempt.attempted_at
                <= datetime.datetime.fromisoformat(end_date)
            )
        return q.all()

    # STRAT-FIX: Redis Caching for Batch Reports (Section 5.6)
    redis_key = f"batch_report:{batch_id}:{start_date}:{end_date}"
    try:
        cached = await redis_client.get(redis_key)
        if cached:
            return json.loads(cached)
    except Exception as e:
        logger.warning(f"Redis cache lookup failed for batch_report: {e}")
        pass

    attempts = await db.run_sync(_load_attempts)
    total_attempts = len(attempts)
    total_members = (
        await db.run_sync(lambda s: s.query(models.User).filter(models.User.group_id.in_(group_ids)).count())
    )

    avg_accuracy = 0.0
    if total_attempts > 0:
        total_score = sum((a.score or 0) for a, _ in attempts)
        total_points = sum((a.total or 0) for a, _ in attempts)
        avg_accuracy = (
            round((total_score / total_points * 100), 2) if total_points > 0 else 0.0
        )

    # Build per-user stats for leaderboard with attempt count
    user_stats: dict = {}
    for a, u in attempts:
        uid = u.id
        if uid not in user_stats:
            user_stats[uid] = {
                "full_name": u.full_name,
                "group_name": u.group.name if u.group else "Member",
                "scores": [],
                "attempt_count": 0,
            }
        acc = (a.score / a.total * 100) if (a.total and a.total > 0) else 0
        user_stats[uid]["scores"].append(acc)
        user_stats[uid]["attempt_count"] += 1

    top_performers = sorted(
        [
            {
                "full_name": s["full_name"],
                "group_name": s["group_name"],
                "avg_score": round(sum(s["scores"]) / len(s["scores"]), 1)
                if s["scores"]
                else 0,
                "attempt_count": s["attempt_count"],
            }
            for s in user_stats.values()
        ],
        key=lambda x: x["avg_score"],
        reverse=True,
    )[:5]

    # Per-group breakdown — avoid N+1 by building a map first
    group_attempt_map: dict = {g.id: [] for g in groups}
    for a, u in attempts:
        if u.group_id in group_attempt_map:
            group_attempt_map[u.group_id].append(a)

    group_breakdown = []
    for g in groups:
        g_atts = group_attempt_map[g.id]
        g_total = len(g_atts)
        g_avg_acc = 0.0
        if g_total > 0:
            g_score = sum((a.score or 0) for a in g_atts)
            g_points = sum((a.total or 0) for a in g_atts)
            g_avg_acc = round((g_score / g_points * 100), 2) if g_points > 0 else 0.0
        group_breakdown.append(
            {
                "id": g.id,
                "group_name": g.name,
                "attempts": g_total,
                "avg_score": g_avg_acc,
            }
        )

    # Generate Strategic Observations using AI
    stats_for_ai = {
        "average_score": avg_accuracy,
        "total_members": total_members,
        "total_attempts": total_attempts,
        "group_performance": group_breakdown,
        "top_performers": top_performers,
    }
    vertical_name = await db.run_sync(
        lambda sync_db: (
            sync_db.query(models.Vertical.name)
            .filter(models.Vertical.id == batch.vertical_id)
            .scalar()
        )
        or "N/A"
    )

    observations = await ai_executive.generate_batch_insights(batch.name, stats_for_ai)

    report = {
        "batch_name": batch.name,
        # `batch.vertical` is a lazy relationship; an AsyncSession cannot resolve
        # it implicitly (MissingGreenlet). Resolved explicitly above.
        "vertical_name": vertical_name,
        "total_members": total_members,
        "total_groups": len(groups),
        "total_attempts": total_attempts,
        "average_score": avg_accuracy,
        "top_performers": top_performers,
        "group_performance": group_breakdown,
        "strategic_observations": observations,
        "from_cache": False,
    }

    try:
        await redis_client.set(redis_key, json.dumps(report), ex=21600)  # 6h TTL
    except Exception as e:
        logger.warning(f"Redis cache lookup failed for group_report: {e}")
        pass

    return report

@router.get("/batch/{batch_id}/xlsx")
async def export_batch_xlsx(
    batch_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(require_ldadmin),
):
    """Professional Multi-Sheet L&D Executive Export."""
    await db.run_sync(lambda s: assert_batch_in_org(batch_id, s, current_user))
    from cache_manager import redis_client

    lock_key = f"rl:export_batch_xlsx:{current_user['sub']}"
    try:
        acquired = await redis_client.set(lock_key, "locked", ex=30)
        if not acquired:
            raise HTTPException(
                status_code=429,
                detail="Export already in progress or requested too recently. Please wait.",
            )
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        pass

    import io

    from fastapi.responses import StreamingResponse

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    batch = await db.run_sync(lambda s: s.query(models.Batch).filter(models.Batch.id == batch_id).first())
    if not batch:
        raise HTTPException(status_code=404)

    groups = await db.run_sync(lambda s: s.query(models.Group).filter(models.Group.batch_id == batch_id).all())
    group_ids = [g.id for g in groups]

    # Fetch data and generate AI insights first
    report_data = await get_batch_report(
        batch_id=batch_id, db=db, current_user=current_user
    )
    rows = (
        await db.run_sync(lambda s: s.query(models.Attempt, models.User)
        .join(models.User)
        .filter(models.User.group_id.in_(group_ids))
        .all())
    )

    wb = openpyxl.Workbook()
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF")
    Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # --- Sheet 1: Executive Summary ---
    ws_summary = wb.active
    assert ws_summary is not None
    ws_summary.title = "Executive Summary"

    ws_summary.merge_cells("A1:E1")
    ws_summary["A1"] = f"STRATEGIC REPORT: {batch.name.upper()}"
    ws_summary["A1"].font = Font(bold=True, size=16, color="3730A3")
    ws_summary["A1"].alignment = Alignment(horizontal="center")

    ws_summary.append([])
    ws_summary.append(["METRIC", "CURRENT STATUS"])
    for cell in ws_summary[3]:
        cell.font = header_font
        cell.fill = header_fill

    ws_summary.append(["Global Proficiency", f"{report_data['average_score']}%"])
    ws_summary.append(["Total Active Members", report_data["total_members"]])
    ws_summary.append(["Engagement Index (Attempts)", report_data["total_attempts"]])
    ws_summary.append(
        [
            "Report Iteration",
            datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%d %H:%M"),
        ]
    )

    ws_summary.append([])
    ws_summary.append(["AI STRATEGIC OBSERVATIONS"])
    ws_summary.merge_cells(f"A{ws_summary.max_row}:E{ws_summary.max_row}")
    ws_summary[f"A{ws_summary.max_row}"].font = Font(bold=True, color="FFFFFF")
    ws_summary[f"A{ws_summary.max_row}"].fill = PatternFill("solid", fgColor="4338CA")

    for obs in report_data.get("strategic_observations", []):
        ws_summary.append([f"• {obs}"])
        ws_summary.merge_cells(f"A{ws_summary.max_row}:E{ws_summary.max_row}")

    # --- Sheet 2: Group Performance ---
    ws_groups = wb.create_sheet("Group Performance")
    ws_groups.append(
        ["GROUP NAME", "ATTEMPTS", "AVG PROFICIENCY %", "ENGAGEMENT LEVEL"]
    )
    for cell in ws_groups[1]:
        cell.font = header_font
        cell.fill = header_fill

    for g in report_data["group_performance"]:
        status = (
            "EXCEPTIONAL"
            if g["avg_score"] > 85
            else "OPTIMAL"
            if g["avg_score"] > 70
            else "NEEDS SYNC"
        )
        ws_groups.append([g["group_name"], g["attempts"], f"{g['avg_score']}%", status])

    # --- Sheet 3: Individual Leaderboard ---
    ws_members = wb.create_sheet("Member Registry")
    ws_members.append(
        ["MEMBER NAME", "GROUP", "TOTAL ATTEMPTS", "AVG ACCURACY %", "TOP SCORE"]
    )
    for cell in ws_members[1]:
        cell.font = header_font
        cell.fill = header_fill

    for p in report_data["top_performers"]:
        ws_members.append(
            [
                p["full_name"],
                p["group_name"],
                p["attempt_count"],
                f"{p['avg_score']}%",
                "-",
            ]
        )

    # --- Sheet 4: Raw Logs ---
    ws_raw = wb.create_sheet("Raw Activity Logs")
    headers = [
        "Attempt ID",
        "User ID",
        "Full Name",
        "Group",
        "Score",
        "Total",
        "Accuracy %",
        "Timestamp",
    ]
    ws_raw.append(headers)
    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill

    for a, u in rows:
        acc = round((a.score / a.total * 100), 1) if a.total > 0 else 0
        ws_raw.append(
            [
                a.id,
                u.id,
                u.full_name,
                u.group.name if u.group else "N/A",
                a.score,
                a.total,
                acc,
                a.attempted_at.strftime("%Y-%m-%d %H:%M"),
            ]
        )

    from openpyxl.utils import get_column_letter

    for sheet in wb.worksheets:
        for col in sheet.columns:
            if not col:
                continue
            max_length = 0
            column_letter = get_column_letter(col[0].column)  # type: ignore
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception as e:
                    logger.warning(f"Column width calculation failed: {e}")
                    pass
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=batch_{batch_id}_executive_report.xlsx"
        },
    )

@router.get("/compare")
def compare_batches(
    batch_ids: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_ldadmin),
):
    """PHASE-3: Side-by-side benchmarking for multiple batches."""
    ids = [int(i.strip()) for i in batch_ids.split(",") if i.strip().isdigit()]
    result = []
    for bid in ids:
        batch = db.query(models.Batch).filter(models.Batch.id == bid).first()
        if not batch:
            continue

        g_ids = [
            g.id
            for g in db.query(models.Group).filter(models.Group.batch_id == bid).all()
        ]

        # Quiz Stats
        attempts = (
            db.query(models.Attempt)
            .join(models.User)
            .filter(models.User.group_id.in_(g_ids))
            .all()
        )
        q_acc = (
            sum(a.score for a in attempts) / sum(a.total for a in attempts) * 100
            if attempts and sum(a.total for a in attempts) > 0
            else 0
        )

        # Coding Stats
        c_atts = (
            db.query(models.CodingAttempt)
            .join(models.User)
            .filter(models.User.group_id.in_(g_ids))
            .all()
        )
        c_acc = (
            sum(ca.score for ca in c_atts if ca.score)
            / len([ca for ca in c_atts if ca.score])
            if c_atts and len([ca for ca in c_atts if ca.score]) > 0
            else 0
        )

        result.append(
            {
                "batch_id": bid,
                "batch_name": batch.name,
                "quiz_accuracy": round(q_acc, 1),
                "coding_proficiency": round(c_acc, 1),
                "active_members": db.query(models.User)
                .filter(models.User.group_id.in_(g_ids))
                .count(),
                "engagement_score": round(
                    (len(attempts) + len(c_atts)) / max(1, len(g_ids)), 1
                ),
            }
        )
    return result

@router.get("/batch/{batch_id}/export")
async def export_batch_report(
    batch_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: dict = Depends(require_ldadmin),
):
    """FUNC-006: Professional multi-sheet Excel export for L&D Stakeholders."""
    await db.run_sync(lambda s: assert_batch_in_org(batch_id, s, current_user))
    batch = await db.run_sync(lambda s: s.query(models.Batch).filter(models.Batch.id == batch_id).first())
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")

    # 1. Gather Data
    groups = await db.run_sync(lambda s: s.query(models.Group).filter(models.Group.batch_id == batch_id).all())
    group_ids = [g.id for g in groups]
    users = (
        await db.run_sync(lambda s: s.query(models.User)
        .filter(models.User.group_id.in_(group_ids), models.User.role == "Member")
        .all())
    )
    user_ids = [u.id for u in users]

    quiz_attempts = (
        await db.run_sync(lambda s: s.query(models.Attempt).filter(models.Attempt.user_id.in_(user_ids)).all())
    )
    coding_attempts = (
        await db.run_sync(lambda s: s.query(models.CodingAttempt)
        .filter(models.CodingAttempt.user_id.in_(user_ids))
        .all())
    )

    # 2. Setup Workbook
    wb = Workbook()

    # --- Sheet 1: Executive Insights ---
    ws1 = wb.active  # type: ignore
    assert ws1 is not None
    ws1.title = "Executive Insights"
    ws1.append(["GO-LIVE STRATEGIC ANALYSIS", "", "GRINDBUDDY V3 L&D ECOSYSTEM"])
    ws1.append(
        [
            f"Batch: {batch.name}",
            "",
            f"Generated: {datetime.datetime.now(datetime.timezone.utc).strftime('%Y-%m-%d %H:%M')}",
        ]
    )
    ws1.append([])

    # Header styling
    for cell in ws1[1]:
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="4F46E5", end_color="4F46E5", fill_type="solid"
        )

    # Fetch AI Summary
    summary_data = {
        "groups": [g.name for g in groups],
        "user_count": len(users),
        "quiz_attempts": len(quiz_attempts),
        "coding_attempts": len(coding_attempts),
        "avg_quiz_score": sum(a.score for a in quiz_attempts) / len(quiz_attempts)
        if quiz_attempts
        else 0,
    }
    ai_bullets = await ai_executive.generate_batch_executive_summary(
        batch.name, summary_data
    )

    ws1.append(["AI-Generated Pedagogical Strategy:"])
    ai_data = ai_bullets.get("data", "")
    for i, line in enumerate(ai_data.split("\n")):
        if line.strip():
            ws1.append([line.strip()])
            ws1.cell(ws1.max_row, 1).alignment = Alignment(wrap_text=True)

    # --- Sheet 2: Performance Registry ---
    ws2 = wb.create_sheet("Performance Registry")
    ws2.append(
        [
            "User ID",
            "Full Name",
            "Group",
            "Department",
            "Avg Quiz Score (%)",
            "Avg Coding Score (10)",
            "Total Attempts",
        ]
    )

    user_map = {u.id: u for u in users}
    group_map = {g.id: g.name for g in groups}

    for uid, u in user_map.items():
        u_quizzes = [a for a in quiz_attempts if a.user_id == uid]
        u_coding = [a for a in coding_attempts if a.user_id == uid]

        avg_q = (
            (sum(a.score / a.total * 100 for a in u_quizzes) / len(u_quizzes))
            if u_quizzes
            else 0
        )
        avg_c = (sum(getattr(a, "score") or 0 for a in u_coding) / len(u_coding)) if u_coding else 0

        ws2.append(
            [
                uid,
                u.full_name,
                group_map.get(u.group_id, "N/A"),
                u.department_id,
                round(avg_q, 1),
                round(avg_c, 1),
                len(u_quizzes) + len(u_coding),
            ]
        )

    # --- Sheet 3: Raw Activity Log ---
    ws3 = wb.create_sheet("Raw Activity Log")
    ws3.append(
        ["Timestamp", "User", "Type", "Activity Name", "Score", "Total", "Status"]
    )

    # Combined logs
    logs = []
    for a in quiz_attempts:
        bank = (
            await db.run_sync(lambda s: s.query(models.QuestionBank)
            .filter(models.QuestionBank.id == a.bank_id)
            .first())
        )
        logs.append(
            [
                a.attempted_at,
                a.user_name,
                "Quiz",
                bank.name if bank else "Quiz",
                a.score,
                a.total,
                "N/A",
            ]
        )
    for a in coding_attempts:
        q = (
            await db.run_sync(lambda s: s.query(models.CodingQuestion)
            .filter(models.CodingQuestion.id == a.coding_question_id)
            .first())
        )
        logs.append(
            [
                a.attempted_at,
                user_map[a.user_id].full_name if a.user_id in user_map else "Unknown",
                "Coding",
                q.title if q else "Lab",
                a.score,
                10,
                "Verified" if a.is_verified else "Pending",
            ]
        )

    for log in sorted(logs, key=lambda x: x[0], reverse=True):
        ws3.append(log)

    # 3. Finalize
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from fastapi.responses import StreamingResponse

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=GrindBuddy_Batch_{batch_id}_Report.xlsx"
        },
    )

@router.get("/batch/{batch_id}/csv")
def export_batch_csv(
    batch_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_ldadmin),
):
    """CSV export for batch performance data — lighter alternative to Excel."""
    assert_batch_in_org(batch_id, db, current_user)
    import csv
    import io as _io

    from fastapi.responses import StreamingResponse as SR

    batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")

    groups = db.query(models.Group).filter(models.Group.batch_id == batch_id).all()
    group_ids = [g.id for g in groups]
    users = db.query(models.User).filter(models.User.group_id.in_(group_ids)).all()
    user_ids = [u.id for u in users]

    quiz_attempts = (
        db.query(models.Attempt).filter(models.Attempt.user_id.in_(user_ids)).all()
    )
    coding_attempts = (
        db.query(models.CodingAttempt)
        .filter(models.CodingAttempt.user_id.in_(user_ids))
        .all()
    )

    group_map = {g.id: g.name for g in groups}

    output = _io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "User ID",
            "Full Name",
            "Group",
            "Email",
            "Avg Quiz Score (%)",
            "Avg Coding Score",
            "Total Attempts",
            "Coding Attempts",
        ]
    )

    for u in users:
        u_quiz = [a for a in quiz_attempts if a.user_id == u.id]
        u_code = [a for a in coding_attempts if a.user_id == u.id]
        avg_q = (
            round(sum(a.score / a.total * 100 for a in u_quiz) / len(u_quiz), 1)
            if u_quiz
            else 0
        )
        avg_c = round(sum(getattr(a, "score") or 0 for a in u_code) / len(u_code), 1) if u_code else 0
        writer.writerow(
            [
                u.id,
                u.full_name,
                group_map.get(u.group_id, "N/A"),
                u.email or "",
                avg_q,
                avg_c,
                len(u_quiz),
                len(u_code),
            ]
        )

    output.seek(0)
    return SR(
        iter([output.getvalue().encode()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=GrindBuddy_Batch_{batch_id}.csv"
        },
    )
